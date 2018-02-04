from openpyxl import Workbook
from openpyxl import load_workbook
from dateutil.parser import parse


def write_purchases(purchases, xls):
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active
    tot_row = 2
    ws.title = "All items"
    ws['A1'] = "Count"
    ws['B1'] = "Item"
    ws['C1'] = "Category"
    ws['D1'] = "Unit Price"
    ws['E1'] = "Estimated Price"
    ws['F1'] = "Total:"
    seen = {}
    dates = list(purchases.keys())
    for date in dates:
        ws1 = wb.create_sheet(title=date.strftime("%Y-%m-%d"))
        row = 1
        for item in purchases[date]:
            # order.append({'name': name, 'link': link, 'count': count, 'unit_price': unit_price})
            cell = ws1.cell(column=1, row=row, value=item['name'])
            cell.hyperlink = item['link']
            c2 = ws1.cell(column=2, row=row, value=item['count'])
            c3 = ws1.cell(column=3, row=row, value=item['unit_price'])
            ws1.cell(column=4, row=row, value='={}*{}'.format(c2.coordinate, c3.coordinate))
            row += 1
            if item['link'] not in seen:
                seen[item['link']] = 1
                cell = ws.cell(column=2, row=tot_row, value=item['name'])
                cell.hyperlink = item['link']
                c2 = ws.cell(column=4, row=tot_row, value=item['unit_price'])
                ws.cell(column=5, row=tot_row, value="=A{}*{}".format(tot_row,c2.coordinate))
                tot_row += 1

    # Save the file
    ws['G1'] = "=SUM(D2:D{})".format(tot_row)
    wb.save(xls)


def build_order(xls):
    wb = load_workbook(filename=xls)
    ws = wb['All items']
    items = {}
    for row in range(2, ws.max_row):
        count = ws.cell(row, 1).value
        if not count:
            continue
        link = ws.cell(row, 2).hyperlink.target
        items[link] = count
    return items

#items = create_order('export.xlsx')
#cache = load_cache()
#purchases = {}
#for date in cache:
#    d = parse(date)   
#    purchases[d] = cache[date]
#write_purchases(purchases, '/media/media/test.xlsx')

